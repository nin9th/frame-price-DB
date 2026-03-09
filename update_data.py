"""
update_data.py
──────────────
Run this whenever you update Frame_Price_DB.xlsx.
It reads all 11 wood types and outputs prices.json.

Usage:
    python update_data.py

Then share prices.json with your teammate via LINE.
"""

import os, sys, shutil, subprocess, json
from datetime import date

EXCEL_PATH  = "Frame_Price_DB.xlsx"
OUTPUT_PATH = "prices.json"
WOOD_TYPES  = [
    "TT-1", "TT-2", "TT-3", "TT-4", "TT-5",
    "TT-6/8", "TT-7/9", "TT-10", "TT-97", "TT-98", "TT-99"
]

try:
    import openpyxl
except ImportError:
    print("❌  openpyxl not found. Run:  pip install openpyxl")
    sys.exit(1)

def recalc(path):
    """Recalculate Excel formulas using LibreOffice or scripts/recalc.py."""
    script = os.path.join(os.path.dirname(__file__), "scripts", "recalc.py")
    if os.path.exists(script):
        try:
            subprocess.run(["python3", script, path, "30"],
                           capture_output=True, timeout=60)
            return
        except Exception:
            pass
    for cmd in ["libreoffice", "libreoffice7.6", "soffice"]:
        if shutil.which(cmd):
            try:
                subprocess.run(
                    [cmd, "--headless", "--calc",
                     "--convert-to", "xlsx",
                     "--outdir", os.path.dirname(os.path.abspath(path)), path],
                    capture_output=True, timeout=60
                )
            except Exception:
                pass
            return

def extract(excel_path, wood_type):
    tmp = f"_tmp_{wood_type.replace('/', '_')}.xlsx"
    shutil.copy(excel_path, tmp)

    wb = openpyxl.load_workbook(tmp)
    wb["Sheet1"]["B11"] = wood_type
    wb.save(tmp)
    wb.close()

    recalc(tmp)

    wb2 = openpyxl.load_workbook(tmp, data_only=True)
    ws  = wb2["ใบราคา"]

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    data = {
        "description":   str(cell(1, 2) or ""),
        "perInchPlain":  round(float(cell(3, 2)), 2) if cell(3, 2) else 0,
        "perInchGroove": round(float(cell(3, 5)), 2) if cell(3, 5) else 0,
        "grooveLabel":   str(cell(6, 5) or "ใส่ฝ้าย 1 นิ้ว"),
        "sizes": []
    }

    for r in range(9, 27):
        size  = cell(r, 1)
        plain = cell(r, 2)
        groove= cell(r, 5)
        if size and plain:
            data["sizes"].append({
                "size":   str(size),
                "plain":  round(float(plain), 1),
                "groove": round(float(groove), 1) if groove else None
            })

    wb2.close()
    os.remove(tmp)
    return data

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌  ไม่พบไฟล์: {EXCEL_PATH}")
        print(f"    วางไฟล์ Frame_Price_DB.xlsx ในโฟลเดอร์เดียวกับสคริปต์นี้")
        sys.exit(1)

    print(f"📂  อ่านข้อมูลจาก: {EXCEL_PATH}\n")
    all_data = {}

    for wood in WOOD_TYPES:
        print(f"   ⚙️  {wood}...", end=" ", flush=True)
        try:
            all_data[wood] = extract(EXCEL_PATH, wood)
            print(f"✓  ({len(all_data[wood]['sizes'])} ขนาด)")
        except Exception as e:
            print(f"❌  {e}")

    output = {
        "updated": date.today().isoformat(),
        "woods": all_data
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"\n✅  สำเร็จ! สร้างไฟล์ {OUTPUT_PATH} ({size_kb:.1f} KB)")
    print(f"    ส่งไฟล์นี้ให้ทีมงานผ่าน LINE ได้เลย 📤")

if __name__ == "__main__":
    main()
